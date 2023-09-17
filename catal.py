from openpyxl import *
from matplotlib import pyplot as plt
from scipy.interpolate import make_interp_spline
import numpy as np
from tkinter import filedialog
from tkinter import *
from tkinter import messagebox as mb
import customtkinter

customtkinter.set_appearance_mode("Dark")
customtkinter.set_default_color_theme("blue")

root = customtkinter.CTk()
root.title('Catal')
root.geometry('300x120')
root.resizable(height=False, width=False)

"""
Для установки библиотек напишите в терминале: pip install matplotlib
                                              pip install numpy
                                              pip install scipy
                                              pip install openpyxl
                                              pip install tkinter
"""


# Название файла, если находится вместе с папкой программы, либо требуется указать путь
def open_file():
    my_file = filedialog.askopenfilename()
    wb = load_workbook(my_file)
    global ws
    ws = wb.active


def reset_lists():
    global C2H2_mm, C2H4_mm, CO_mm, H2_mm, F_eef_mm, F_vsg_mm, T_vh_mm
    global F_eef, F_vsg, C2H2_1, C2H2_2, C2H4_1, C2H4_2, CO_1, CO_2, H2_1, H2_then, H2_2, T_vh
    global F_eef_n, F_vsg_n, C2H2_1_n, C2H2_2_n, C2H4_1_n, C2H4_2_n, CO_1_n, CO_2_n, H2_1_n, H2_then_n, H2_2_n, T_vh_n
    global F_eef_m, F_vsg_m, C2H2_1_m, C2H2_2_m, C2H4_1_m, C2H4_2_m, CO_1_m, CO_2_m, H2_1_m, H2_then_m, H2_2_m, T_vh_m
    global F_eef_min, F_vsg_min, C2H2_1_min, C2H2_2_min, C2H4_1_min, C2H4_2_min, CO_1_min, CO_2_min, H2_1_min, H2_then_min, H2_2_min, T_vh_min
    global el_name, normal, el_name_n, table_name, ga_array, mid_count, min_normal
    global months, data_array, cycle_count, array_numbers_for_value, data_ga, col_sum, col_sum_min_max, col_sum_normal, col_sum_normal_minus
    global numerated, numerated2, numerated3, final, mid_count_csnm, csnm_value

    # Очистка списков
    C2H2_mm = []
    C2H4_mm = []
    CO_mm = []
    H2_mm = []
    F_eef_mm = []
    F_vsg_mm = []
    T_vh_mm = []

    F_eef = []
    F_vsg = []
    C2H2_1 = []
    C2H2_2 = []
    C2H4_1 = []
    C2H4_2 = []
    CO_1 = []
    CO_2 = []
    H2_1 = []
    H2_then = []
    H2_2 = []
    T_vh = []

    F_eef_n = []
    F_vsg_n = []
    C2H2_1_n = []
    C2H2_2_n = []
    C2H4_1_n = []
    C2H4_2_n = []
    CO_1_n = []
    CO_2_n = []
    H2_1_n = []
    H2_then_n = []
    H2_2_n = []
    T_vh_n = []

    F_eef_m = []
    F_vsg_m = []
    C2H2_1_m = []
    C2H2_2_m = []
    C2H4_1_m = []
    C2H4_2_m = []
    CO_1_m = []
    CO_2_m = []
    H2_1_m = []
    H2_then_m = []
    H2_2_m = []
    T_vh_m = []

    F_eef_min = []
    F_vsg_min = []
    C2H2_1_min = []
    C2H2_2_min = []
    C2H4_1_min = []
    C2H4_2_min = []
    CO_1_min = []
    CO_2_min = []
    H2_1_min = []
    H2_then_min = []
    H2_2_min = []
    T_vh_min = []

    # Заполнение списков с изначальными значениями
    el_name = [F_eef, F_vsg, C2H2_1, C2H2_2, C2H4_1, C2H4_2, CO_1, CO_2, H2_1, H2_then, H2_2, T_vh]

    # Заполнение списков с нормализованными значениями
    normal = [F_eef_n, F_vsg_n, C2H2_1_n, C2H2_2_n, C2H4_1_n, C2H4_2_n, CO_1_n,
              CO_2_n, H2_1_n, H2_then_n, H2_2_n, T_vh_n]

    # Заполнение списка с наименованиями значений
    el_name_n = ['F_eef', 'F_vsg', 'C2H2_1', 'C2H2_2', 'C2H4_1', 'C2H4_2',
                 'CO_1', 'CO_2', 'H2_1', 'H2_then', 'H2_2', 'T_vh']

    # Упорядоченный список в соответствии со значениями
    table_name = ['B', 'C', 'O', 'T', 'P', 'U',
                  'Q', 'V', 'R', 'S', 'W', 'D']

    # Значения ген алгоритма
    ga_array = [3.008, -1.534, 5.74683, 2.00123, -0.6652, 0.02138,
                1.93448, 4.00269, 1.07441, 2.31127, 4.93051, 4.92008]

    # Средние значения
    mid_count = [F_eef_m, F_vsg_m, C2H2_1_m, C2H2_2_m, C2H4_1_m, C2H4_2_m, CO_1_m,
                 CO_2_m, H2_1_m, H2_then_m, H2_2_m, T_vh_m]

    # Создание списка с меньшим количеством нормализованных значений
    min_normal = [F_eef_min, F_vsg_min, C2H2_1_min, C2H2_2_min, C2H4_1_min, C2H4_2_min, CO_1_min,
                  CO_2_min, H2_1_min, H2_then_min, H2_2_min, T_vh_min]

    # Вспомогательные списки
    months = []
    data_array = []
    cycle_count = [0]
    array_numbers_for_value = []
    data_ga = [[], [], [], [], [], [], [], [], [], [], [], []]
    col_sum = []
    col_sum_min_max = []
    col_sum_normal = []
    col_sum_normal_minus = []
    numerated = []
    numerated2 = []
    numerated3 = []
    final = []
    mid_count_csnm = []
    csnm_value = []


# Добавьте вызов функции reset_lists перед началом расчетов:

def start():
    try:
        reset_lists()  # Очистка и возврат списков к начальному состоянию
        all_func()
    except NameError:
        mb.showerror(
            'Ошибка!',
            'Для запуска нужно выбрать файл!')

K_CONST = 5


# Создание массивов с изначальными данными
def cr_array(t_name, e_name):
    for i in range(len(t_name)):
        for row in range(1, ws.max_row + 1):
            e_name[i].append(float(ws[t_name[i] + str(row)].value))


# Создание минимальных и максимальных значений
def min_max(e_name):
    for i in range(len(e_name)):
        e_name[i].sort()
    for i in range(len(e_name)):
        if e_name[i] == C2H2_1 or e_name[i] == C2H2_2:
            C2H2_mm.append(e_name[i][0])
            C2H2_mm.append(e_name[i][-1])
            C2H2_mm.sort()
        elif e_name[i] == C2H4_1 or e_name[i] == C2H4_2:
            C2H4_mm.append(e_name[i][0])
            C2H4_mm.append(e_name[i][-1])
            C2H4_mm.sort()
        elif e_name[i] == CO_1 or e_name[i] == CO_2:
            CO_mm.append(e_name[i][0])
            CO_mm.append(e_name[i][-1])
            CO_mm.sort()
        elif e_name[i] == H2_1 or e_name[i] == H2_2 or e_name[i] == H2_then:
            H2_mm.append(e_name[i][0])
            H2_mm.append(e_name[i][-1])
            H2_mm.sort()
        elif e_name[i] == F_eef:
            F_eef_mm.append(e_name[i][0])
            F_eef_mm.append(e_name[i][-1])
            F_eef_mm.sort()
        elif e_name[i] == F_vsg:
            F_vsg_mm.append(e_name[i][0])
            F_vsg_mm.append(e_name[i][-1])
            F_vsg_mm.sort()
        elif e_name[i] == T_vh:
            T_vh_mm.append(e_name[i][0])
            T_vh_mm.append(e_name[i][-1])
            T_vh_mm.sort()
    for i in range(len(e_name)):
        e_name[i].clear()
    cr_array(table_name, el_name)


# Нормализация списков
def norm_cr_array(norm, e_name):
    for j in range(len(norm)):
        for i in range(len(F_eef)):
            if norm[j] == F_eef_n:
                norm[j].append(0 + ((e_name[j][i]) - (F_eef_mm[0])) / ((F_eef_mm[-1]) - (F_eef_mm[0])) * (1 - 0))
            elif norm[j] == F_vsg_n:
                norm[j].append(0 + ((e_name[j][i]) - (F_vsg_mm[0])) / ((F_vsg_mm[-1]) - (F_vsg_mm[0])) * (1 - 0))
            elif norm[j] == C2H2_1_n or norm[j] == C2H2_2_n:
                norm[j].append(0 + ((e_name[j][i]) - (C2H2_mm[0])) / ((C2H2_mm[-1]) - (C2H2_mm[0])) * (1 - 0))
            elif norm[j] == C2H4_1_n or norm[j] == C2H4_2_n:
                norm[j].append(0 + ((e_name[j][i]) - (C2H4_mm[0])) / ((C2H4_mm[-1]) - (C2H4_mm[0])) * (1 - 0))
            elif norm[j] == CO_1_n or norm[j] == CO_2_n:
                norm[j].append(0 + ((e_name[j][i]) - (CO_mm[0])) / ((CO_mm[-1]) - (CO_mm[0])) * (1 - 0))
            elif norm[j] == H2_1_n or norm[j] == H2_2_n or norm[j] == H2_then_n:
                norm[j].append(0 + ((e_name[j][i]) - (H2_mm[0])) / ((H2_mm[-1]) - (H2_mm[0])) * (1 - 0))
            elif norm[j] == T_vh_n:
                norm[j].append(0 + ((e_name[j][i]) - (T_vh_mm[0])) / ((T_vh_mm[-1]) - (T_vh_mm[0])) * (1 - 0))
            else:
                break


# Выбор только значений месяца из даты
def only_months(data, mon):
    for row in range(1, ws.max_row + 1):
        data.append(ws['A' + str(row)].value)
    for i in range(len(data)):
        mon.append(int(data[i].strftime("%m")))


# Создание критических значений цикла(конец цикла/полугодия)
def _cycle(mon, cycle):
    for i in range(len(mon)):
        if i + 1 >= len(mon):
            cycle.append(i)
            break
        if 2 < mon[i + 1] - mon[i] < 11:
            cycle.append(i)
        else:
            continue


# Разделение данных на секции
def _sections(cycle, arnfv):
    for i in range(len(cycle)):
        if i + 1 >= len(cycle):
            break
        else:
            k_const_values = ((cycle[i + 1] - cycle[i]) // K_CONST)
        arnfv.append(k_const_values - 1)


# Уменьшение размера нормализации данных с определенным шагом(array_numbers_for_value[_step]
# до критического значения cycle_count[_step+1]
def _minimize_normal(norm, cycle, arnfv, mn):
    for i in range(len(norm)):
        for _step in range(len(cycle) - 1):
            for j in range(cycle[_step], cycle[_step + 1], arnfv[_step]):
                if j < cycle[_step + 1]:
                    mn[i].append(
                        (sum(norm[i][j:j + arnfv[_step]])) / arnfv[_step])
    for i in range(len(norm)):
        for j in range(K_CONST, (K_CONST * (len(cycle))), K_CONST):
            mn[i].pop(j)


# Функция заполнения массива умноженных минимальных данных на данные из ген алгоритма
def data_for_ga(mn, dg, ga):
    for i in range(len(mn)):
        for j in range(len(mn[i])):
            dg[i].append(mn[i][j] * ga[i])


# Сумма значений столбцов матрицы
def matrix_data():
    for i in range(len(data_ga[0])):
        columns = 0
        for j in range(len(data_ga)):
            columns += data_ga[j][i]
        col_sum.append(columns)


# Минимальные и максимальные значения списка суммы столбцов
def min_max_col_sum():
    col_sum.sort()
    col_sum_min_max.append(col_sum[0])
    col_sum_min_max.append(col_sum[-1])
    col_sum.clear()
    matrix_data()


# Нормальные значения суммы столбцов матрицы и построение графика
def normal_col_sum():
    numerated.clear()
    for i in range(len(col_sum)):
        col_sum_normal.append(
            ((0 + (col_sum[i]) - (col_sum_min_max[0])) / ((col_sum_min_max[-1]) - col_sum_min_max[0])) * 1)
    for i in range(len(col_sum_normal)):
        col_sum_normal_minus.append(1 - col_sum_normal[i])
    for i in range(len(cycle_count) - 1):
        for j in range(len(col_sum_normal)):
            if j <= int((cycle_count[i + 1] - cycle_count[i]) / array_numbers_for_value[i]):
                numerated.append(cycle_count[i] + (j + 1) * array_numbers_for_value[i])
            else:
                continue
        numerated.pop(-1)


# Создание промежутка для удаления скачков
def _csnm():
    for i in range(len(col_sum_normal_minus) - 1):
        if col_sum_normal_minus[i] - col_sum_normal_minus[i + 1] < 0:
            mid_count_csnm.append(col_sum_normal_minus[i] - col_sum_normal_minus[i + 1])
    csnm_value.append(sum(mid_count_csnm) / len(mid_count_csnm))


def final_array():
    for i in range(len(col_sum_normal_minus) - 1):
        if (col_sum_normal_minus[i] - col_sum_normal_minus[i + 1] < csnm_value[0]) \
                or (col_sum_normal_minus[i] - col_sum_normal_minus[i + 1] > 0):
            final.append(col_sum_normal_minus[i])
            numerated3.append(col_sum_normal_minus.index(col_sum_normal_minus[i]))
    for i in range(len(final)):
        numerated2.append(numerated[numerated3[i]])


# Сглаживаение графика
def interpolate(x_array, y_array):
    x = np.array(x_array)
    y = np.array(y_array)
    xnew = np.linspace(x.min(), x.max(), 200)
    spl = make_interp_spline(x, y, k=2)
    y_smooth = spl(xnew)
    plt.plot(xnew, y_smooth)
    plt.show()


def all_func():
    cr_array(table_name, el_name)
    min_max(el_name)
    norm_cr_array(normal, el_name)
    only_months(data_array, months)
    _cycle(months, cycle_count)
    _sections(cycle_count, array_numbers_for_value)
    _minimize_normal(normal, cycle_count, array_numbers_for_value, min_normal)
    data_for_ga(min_normal, data_ga, ga_array)
    matrix_data()
    min_max_col_sum()
    normal_col_sum()
    _csnm()
    final_array()
    interpolate(numerated, col_sum_normal_minus)
    interpolate(numerated2, final)


my_button = customtkinter.CTkButton(root, text="Провести расчеты", command=start)
my_button.pack(pady=35)


def help_menu():
    top = customtkinter.CTkToplevel()
    top.resizable(width=False, height=False)
    top.title('Помощь')
    text_help = customtkinter.CTkTextbox(top, width=300, height=200, wrap=WORD)
    text_help.pack()
    text_help.insert(1.0,
                     "1) Для работы необходимо указать путь к файлу формата xlsx, csv."
                     " \n2) После выбора файла запустите программу нажав на кнопку \"Провести расчеты\"."
                     " \n3) Если файл не будет выбран, то всплывет диалоговое окно с ошибкой."
                     " \n4) В случае выбора правильного файла программа будет выполнена."
                     )


def about_menu():
    top = customtkinter.CTkToplevel()
    top.resizable(width=False, height=False)
    top.title('О программе')
    text_help = customtkinter.CTkTextbox(top, width=500, height=250, wrap=WORD)
    text_help.pack()
    text_help.insert(1.0,
                     "    Программа оценки состояния катализатора использует данные загруженные с "
                     "excel-файла для построения графика, показывающего текущее состояние "
                     "исследуемого объекта."
                     "\n    При наведении курсора на график, можно детально определить значения точек "
                     "лежащих на графике. Также можно взаимодействовать с интерфейсом программы "
                     "используя такие функции как увеличение размера, перемещение области просмотра "
                     "и т.д."
                     "\n    По оси X обозначены дни зафиксированных данных, по оси Y - текущее состояние "
                     "катализатора."
                     "\n    Excel-файл содержит значения следующих параметров (Параметр: столбец): "
                     "\n Fээф: B, Fвсг: C, C2H2(1): O, 2H2(2): T, C2H4(1): P, C2H4(2): U,"
                     " CO(1): Q, CO(2): V, H2(1): R, H2(П): S, H2(2): W, Tвх: D."
                     )


my_menu = Menu(root)
root.config(menu=my_menu)

file_menu = Menu(my_menu, tearoff=0)
my_menu.add_cascade(label="Управление", menu=file_menu)
file_menu.add_command(label="Открыть файл", command=open_file)
file_menu.add_command(label="Помощь", command=help_menu)
file_menu.add_command(label="О программе", command=about_menu)
file_menu.add_separator()
file_menu.add_command(label="Выход", command=root.quit)

root.mainloop()
